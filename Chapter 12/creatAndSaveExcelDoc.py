import openpyxl
# wb = openpyxl.Workbook()

# print(wb.sheetnames)
# sheet=wb.active
# print(sheet.title)
# sheet.title="Spam Bacon Eggs Sheet"
# print(wb.sheetnames)

wb=openpyxl.load_workbook('example.xlsx')
sheet=wb.active 
sheet.title="spam spam spam"
wb.save('example_copy.xlsx')