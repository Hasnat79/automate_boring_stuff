import openpyxl
wb=openpyxl.Workbook()
print(wb.sheetnames)
wb.create_sheet()
print(wb.sheetnames)
wb.create_sheet(index=0,title="First Sheet")
print(wb.sheetnames)
wb.create_sheet(index=2,title="Middle Sheet")
print(wb.sheetnames)

#removing sheets

del wb["Middle Sheet"]
del wb["Sheet1"]
print(wb.sheetnames)