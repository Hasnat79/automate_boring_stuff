import openpyxl
wb=openpyxl.load_workbook('/home/hasnat/Desktop/automate_boring_stuff/Chapter 12/example.xlsx')
# print(type(wb))
# print(wb.sheetnames)
# sheet=wb['Sheet3']
# print(sheet)
# print(type(sheet))
# print(sheet.title)
# anotherSheet=wb.active
# print(anotherSheet)

# getting cells from the sheets
# sheet = wb["Sheet1"]
# print(sheet['A1'])
# print(sheet['A1'].value)
# c = sheet['B1']
# print(c.value)
# print('Row '+str(c.row) + ', column '+str(c.column)+' is '+c.value )
# print('Cell '+c.coordinate + ' is '+c.value )
# print(sheet['C1'].value)

# sheet=wb["Sheet1"]
# print(sheet.cell(row=1,column=2))

# print(sheet.cell(1,2).value )
# for i in range(1,8,2):
#     print(i, sheet.cell(row=i,column=2).value)

sheet=wb['Sheet1']
print(sheet.max_row)
print(sheet.max_column)